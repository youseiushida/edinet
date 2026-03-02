"""生成スクリプトの冪等性テスト。"""
from __future__ import annotations

import csv
import subprocess
import sys
from pathlib import Path

from openpyxl import Workbook


ROOT = Path(__file__).resolve().parents[2]
TOOLS_DIR = ROOT / "tools"


def _run_generator(script_name: str, source: Path, output: Path) -> None:
    cmd = [
        sys.executable,
        str(TOOLS_DIR / script_name),
        "--source",
        str(source),
        "--output",
        str(output),
    ]
    subprocess.run(cmd, check=True, cwd=ROOT)


def _build_form_xlsx(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "様式コードリスト"
    headers = [
        "府令コード",
        "様式コード（コード値）",
        "様式番号",
        "様式名",
        "書類種別",
        "開示区分",
        "備考",
    ]
    for i, header in enumerate(headers, start=2):
        ws.cell(row=3, column=i, value=header)

    ws.cell(row=4, column=2, value="010")
    ws.cell(row=4, column=3, value="010000")
    ws.cell(row=4, column=4, value="第一号様式")
    ws.cell(row=4, column=5, value="有価証券通知書")
    ws.cell(row=4, column=6, value="有価証券通知書")
    ws.cell(row=4, column=7, value="非開示")
    ws.cell(row=4, column=8, value="")

    ws.cell(row=5, column=2, value="010")
    ws.cell(row=5, column=3, value="010001")
    ws.cell(row=5, column=4, value="第一号様式の二")
    ws.cell(row=5, column=5, value="訂正有価証券通知書")
    ws.cell(row=5, column=6, value="有価証券通知書")
    ws.cell(row=5, column=7, value="開示")
    ws.cell(row=5, column=8, value="備考")

    wb.save(path)


def _build_fund_csv(path: Path) -> None:
    with path.open("w", encoding="cp932", newline="") as fp:
        writer = csv.writer(fp)
        writer.writerow(["ダウンロード実行日", "2026年02月15日現在", "件数", "2件"])
        writer.writerow(
            [
                "ファンドコード",
                "証券コード",
                "ファンド名",
                "ファンド名（ヨミ）",
                "特定有価証券区分名",
                "特定期1",
                "特定期2",
                "ＥＤＩＮＥＴコード",
                "発行者名",
            ]
        )
        writer.writerow(
            ["G01003", "13005", "テストファンドA", "テスト", "投資信託", "1月13日", "", "E00004", "発行者A"]
        )
        writer.writerow(
            ["G01004", "", "テストファンドB", "テスト", "投資信託", "", "", "E00005", "発行者B"]
        )


def _build_edinet_csv(path: Path) -> None:
    with path.open("w", encoding="cp932", newline="") as fp:
        writer = csv.writer(fp)
        writer.writerow(["ダウンロード実行日", "2026年02月15日現在", "件数", "2件"])
        writer.writerow(
            [
                "ＥＤＩＮＥＴコード",
                "提出者種別",
                "上場区分",
                "連結の有無",
                "資本金",
                "決算日",
                "提出者名",
                "提出者名（英字）",
                "提出者名（ヨミ）",
                "所在地",
                "提出者業種",
                "証券コード",
                "提出者法人番号",
            ]
        )
        writer.writerow(
            [
                "E00004",
                "内国法人・組合",
                "上場",
                "有",
                "635,401",
                "3月31日",
                "カネコ種苗株式会社",
                "KANEKO SEEDS CO., LTD.",
                "カネコシュビョウ",
                "群馬県前橋市",
                "水産・農林業",
                "13760",
                "1234567890123",
            ]
        )
        writer.writerow(
            [
                "E00005",
                "内国法人・組合",
                "",
                "",
                "",
                "",
                "テスト株式会社",
                "",
                "",
                "",
                "",
                "",
                "",
            ]
        )


def test_generate_form_codes_idempotent(tmp_path: Path) -> None:
    source = tmp_path / "ESE140327.xlsx"
    output = tmp_path / "form_code.py"
    _build_form_xlsx(source)

    _run_generator("generate_form_codes.py", source, output)
    first = output.read_text(encoding="utf-8")
    _run_generator("generate_form_codes.py", source, output)
    second = output.read_text(encoding="utf-8")
    assert first == second


def test_generate_fund_codes_idempotent(tmp_path: Path) -> None:
    source = tmp_path / "FundcodeDlInfo.csv"
    output = tmp_path / "fund_code.py"
    _build_fund_csv(source)

    _run_generator("generate_fund_codes.py", source, output)
    first = output.read_text(encoding="utf-8")
    _run_generator("generate_fund_codes.py", source, output)
    second = output.read_text(encoding="utf-8")
    assert first == second


def test_generate_edinet_codes_idempotent(tmp_path: Path) -> None:
    source = tmp_path / "EdinetcodeDlInfo.csv"
    output = tmp_path / "edinet_code.py"
    _build_edinet_csv(source)

    _run_generator("generate_edinet_codes.py", source, output)
    first = output.read_text(encoding="utf-8")
    _run_generator("generate_edinet_codes.py", source, output)
    second = output.read_text(encoding="utf-8")
    assert first == second
