"""D-3. F14検証: 添付5 ExcelファイルからAccountingStandardsDEIの値を確認

実行方法: uv run docs/QAs/scripts/D-3.verify_f14.py
前提: openpyxl が必要（uv add openpyxl）
出力: AccountingStandardsDEI に関する記述を添付5 Excelから抽出し、
      D-3.a.md の [F14] の引用内容と突合する
"""

from __future__ import annotations

import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl が必要です。 uv add openpyxl を実行してください。")
    sys.exit(1)

XLSX_PATH = Path(
    "docs/仕様書/2026/提出者別タクソノミ作成ガイドライン添付資料/"
    "提出者別タクソノミ作成ガイドライン (R07.11) 添付5 様式ごとのDEI の設定値対応一覧.xlsx"
)


def main() -> None:
    """添付5 Excelから AccountingStandardsDEI と TypeOfCurrentPeriodDEI の記述を抽出する。"""
    if not XLSX_PATH.exists():
        print(f"ERROR: ファイルが見つかりません: {XLSX_PATH}")
        sys.exit(1)

    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)

    print(f"シート数: {len(wb.sheetnames)}")
    print(f"シート名: {wb.sheetnames}")
    print()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"=== シート: {sheet_name} ===")
        print(f"行数: {ws.max_row}, 列数: {ws.max_column}")

        # まず全行をダンプして構造を把握
        for row_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
            cells = [cell.value for cell in row]
            non_empty = [(i + 1, v) for i, v in enumerate(cells) if v is not None]
            if not non_empty:
                continue
            row_text = " ".join(str(v) for _, v in non_empty)
            # AccountingStandards, TypeOfCurrentPeriod, 会計基準, 当会計期間 を含む行
            keywords = ["AccountingStandards", "TypeOfCurrentPeriod", "会計基準",
                        "当会計期間の種類", "JMIS", "Japan GAAP", "US GAAP", "IFRS",
                        "FY", "HY"]
            if any(kw in row_text for kw in keywords):
                print(f"\n  Row {row_idx}:")
                for col_idx, val in non_empty:
                    val_str = str(val)
                    print(f"    Col {col_idx}: {val_str[:500]}")

        # Row 20 と Row 27 を直接表示（D-3.a.md の [F14] と F-1.a.md の [F18][F19] が参照）
        print("\n--- Row 20 の全内容（D-3 F14 / F-1 F18 が参照）---")
        for row_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
            if row_idx == 20:
                for col_idx, cell in enumerate(row, start=1):
                    if cell.value is not None:
                        print(f"  Col {col_idx}: {cell.value}")

        print("\n--- Row 27 の全内容（F-1 F19 が参照）---")
        for row_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
            if row_idx == 27:
                for col_idx, cell in enumerate(row, start=1):
                    if cell.value is not None:
                        print(f"  Col {col_idx}: {cell.value}")

    wb.close()

    print("\n--- 検証結論 ---")
    print("上記の出力を D-3.a.md [F14] および F-1.a.md [F18][F19] の引用内容と突合してください。")
    print("確認ポイント:")
    print("  1. AccountingStandardsDEI の値に 'JMIS' が含まれているか")
    print("  2. 取りうる値が 'Japan GAAP', 'US GAAP', 'IFRS', 'JMIS' + nil の5種か")
    print("  3. TypeOfCurrentPeriodDEI の値が 'FY', 'HY' の2種のみか")


if __name__ == "__main__":
    main()
